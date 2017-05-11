from openpyxl.styles import Font, NamedStyle, PatternFill

import surveys.models

title_style = NamedStyle(name='title')
title_style.font = Font(size=24, color='F5DD5D')
title_style.fill = PatternFill(fill_type='solid', start_color='417690')

subtitle_style = NamedStyle(name='subtitle')
subtitle_style.font = Font(color='FFFFFF')
subtitle_style.fill = PatternFill(fill_type='solid', start_color='79AEC8')

table_header_style = NamedStyle(name='header')
table_header_style.font = Font(color='FFFFFF')
table_header_style.fill = PatternFill(fill_type='solid', start_color='79AEC8')

table_row_style = NamedStyle(name='table-row')
table_row_style.fill = PatternFill(fill_type='solid', start_color='F8F8F8')

date_style = NamedStyle(name='date')
date_style.number_format = 'dd-mm-yyyy h:mm:ss'
date_style.fill = PatternFill(fill_type='solid', start_color='F8F8F8')


def generate_header(sheet, session, subtitle_complement):
    title = sheet.cell(row=1, column=1)
    title.value = f"Sondage: {session.survey.title}"
    title.style = title_style

    sheet.row_dimensions[1].style = title_style

    subtitle = sheet.cell(row=2, column=1)
    subtitle.value = f"Session: {session.date} - {session.title} - {subtitle_complement}"
    subtitle.style = subtitle_style

    sheet.row_dimensions[2].style = subtitle_style


def export_questions_sheet(sheet, session):
    generate_header(sheet, session, "réponses aux questions")

    questions = session.survey.question_set.all()

    row = 4
    for question in questions:
        question_header = sheet.cell(row=row, column=1)
        question_header.value = question.text
        question_header.style = table_header_style

        answers_header = sheet.cell(row=row+1, column=1)
        answers_header.value = "Réponses"
        answers_header.style = table_row_style

        if question.type == surveys.models.Question.UNIQUE:
            for i, choice in enumerate(question.choice_set.all()):
                choice_header = sheet.cell(row=row, column=1+1+i)
                choice_header.value = choice.text
                choice_header.style = table_header_style

                choice_answers = sheet.cell(row=row+1, column=1+1+i)
                choice_answers.value = surveys.models.UniqueAnswer.objects.filter(
                    question=question,
                    choice=choice,
                    session=session
                ).count()
                choice_answers.style = table_row_style

            row += 3

        elif question.type == surveys.models.Question.MULTIPLE:
            for i, choice in enumerate(question.choice_set.all()):
                choice_header = sheet.cell(row=row, column=1+1+i)
                choice_header.value = choice.text
                choice_header.style = table_header_style

                choice_answers = sheet.cell(row=row+1, column=1+1+i)
                choice_answers.value = surveys.models.MultipleAnswer.objects.filter(
                    question=question,
                    choices=choice,
                    session=session
                ).count()
                choice_answers.style = table_row_style

            row += 3

        elif question.type == surveys.models.Question.RATE:
            for i in range(question.scale + 1):
                rate_header = sheet.cell(row=row, column=1+1+i)
                rate_header.value = i
                rate_header.style = table_header_style

                rate_answers = sheet.cell(row=row+1, column=1+1+i)
                rate_answers.value = surveys.models.RateAnswer.objects.filter(
                    question=question,
                    rating=i,
                    session=session
                ).count()
                rate_answers.style = table_row_style

            row += 3

        elif question.type == surveys.models.Question.RANK:
            choices = question.choice_set.all()

            for i, choice in enumerate(choices):
                rank_header = sheet.cell(row=row, column=1+1+i)
                rank_header.value = i + 1
                rank_header.style = table_header_style

                choice_header = sheet.cell(row=row+1+i, column=1)
                choice_header.value = choice.text
                choice_header.style = table_row_style

                for j, _ in enumerate(choices):
                    rank_count = sheet.cell(row=row+1+i, column=1+1+j)
                    rank_count.value = surveys.models.RankAnswer.objects.filter(
                        question=question,
                        choice=choice,
                        rank=j+1,
                        session=session
                    ).count()
                    rank_count.style = table_row_style

            row += choices.count() + 2


def export_comments_sheet(sheet, session):
    generate_header(sheet, session, "commentaires")

    row = 4
    for i, comment in enumerate(session.comment_set.all()):
        date = sheet.cell(row=row+i, column=1)
        date.value = comment.date
        date.style = date_style

        author = sheet.cell(row=row+i, column=2)
        author.value = str(comment.author)
        author.style = table_row_style

        text = sheet.cell(row=row+i, column=3)
        text.value = comment.text
        text.style = table_row_style

    sheet.column_dimensions['A'].width = 18


def export_respondents_sheet(sheet, session):
    generate_header(sheet, session, "interrogés")

    row = 4
    for i, respondent in enumerate(session.respondent_set.all()):
        first_name = sheet.cell(row=row + i, column=1)
        first_name.value = respondent.first_name
        first_name.style = table_row_style

        last_name = sheet.cell(row=row + i, column=2)
        last_name.value = respondent.last_name
        last_name.style = table_row_style

        gender = sheet.cell(row=row + i, column=3)
        gender.value = respondent.get_gender_display()
        gender.style = table_row_style

        age = sheet.cell(row=row + i, column=4)
        age.value = respondent.age
        age.style = table_row_style
