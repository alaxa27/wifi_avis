import os
import statistics

from django.conf import settings
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.validators import MaxValueValidator, MinValueValidator
from django.db import models
from django.utils.translation import ugettext_lazy as _

import openpyxl as xl

import surveys.export


class Survey(models.Model):
    title = models.CharField(verbose_name=_("Title"), max_length=255)

    class Meta:
        verbose_name = _("Survey")

    @property
    def is_current(self):
        return Session.objects.filter(survey=self, is_current=True).exists()
    verbose_name_plural = _("Surveys")

    def __str__(self):
        return self.title


class Session(models.Model):
    title = models.CharField(verbose_name=_("Title"), max_length=255)
    survey = models.ForeignKey(Survey, verbose_name=_("Survey"), null=False, blank=False, on_delete=models.CASCADE)
    date = models.DateTimeField(verbose_name=_("Date"), auto_now=True)
    is_current = models.BooleanField(verbose_name=_("Current"), blank=True, default=False)

    @classmethod
    def reset_current(cls):
        cls.objects.filter(is_current=True).update(is_current=False)

    @classmethod
    def set_current(cls, new_current):
        cls.reset_current()
        new_current.is_current = True
        new_current.save()

    class Meta:
        verbose_name = _("Session")
        verbose_name_plural = _("Sessions")

    def clean(self, *args, **kwargs):
        # There must only be one current session
        if self.is_current:
            current_sessions = Session.objects.filter(is_current=True)
            if self.pk:
                current_sessions = current_sessions.exclude(pk=self.pk)
            if current_sessions.exists():
                raise ValidationError(_("Only one session can be active at time."))

    def export_to_xlsx(self):
        wb = xl.Workbook()

        surveys.export.export_questions_sheet(wb.create_sheet("Résultats"), self)
        surveys.export.export_comments_sheet(wb.create_sheet("Commentaires"), self)
        surveys.export.export_respondents_sheet(wb.create_sheet("Interrogés"), self)

        i = 0
        while os.path.exists(os.path.join(settings.EXPORT_DIR, f"{self} - {i}.xlsx")):
            i += 1

        wb.save(os.path.join(settings.EXPORT_DIR, f"{self} - {i}.xlsx"))

    def __str__(self):
        return self.title


class Question(models.Model):
    text = models.TextField(verbose_name=_("Question"), blank=False)
    image = models.ImageField(verbose_name=_("Image"), blank=True, null=True)
    survey = models.ForeignKey(Survey, verbose_name=_("Survey"), null=False, on_delete=models.CASCADE)

    UNIQUE = 'UNIQ'
    MULTIPLE = 'MULT'
    RANK = 'RANK'
    RATE = 'RATE'

    TYPES = (
        (UNIQUE, _("Unique")),
        (MULTIPLE, _("Multiple")),
        (RANK, _("Rank")),
        (RATE, _("Rate"))
    )

    type = models.CharField(verbose_name=_("Type"), max_length=4, choices=TYPES, default=UNIQUE)

    # RATE only

    SCALE_NORMAL = 'NORMAL'
    SCALE_STARS = 'STARS'

    SCALE_TYPES = (
        (SCALE_NORMAL, _("Normal")),
        (SCALE_STARS, _("Stars"))
    )

    scale_type = models.CharField(verbose_name=_("Rate type"), max_length=6, choices=SCALE_TYPES, default=SCALE_NORMAL)
    scale = models.PositiveIntegerField(verbose_name=_("Scale"), blank=True, null=True)

    is_active = models.BooleanField(verbose_name=_("Active"), blank=True, default=False)

    class Meta:
        verbose_name = _("Question")
        verbose_name_plural = _("Questions")

    def get_answers_data(self):
        if self.type == self.UNIQUE:
            return [
                UniqueAnswer.objects.filter(session__is_current=True, question=self, choice=choice).count()
                for choice in self.choice_set.all()
            ]
        elif self.type == self.MULTIPLE:
            return [
                MultipleAnswer.objects.filter(session__is_current=True, question=self, choices=choice).count()
                for choice in self.choice_set.all()
            ]
        elif self.type == self.RATE:
            return [
                RateAnswer.objects.filter(session__is_current=True, question=self, rating=rating).count()
                for rating in range(self.scale + 1)
            ]
        elif self.type == self.RANK:
            return [
                1 / statistics.mean([a.rank for a in RankAnswer.objects.filter(
                    session__is_current=True,
                    question=self,
                    choice=choice
                )] or [0])
                for choice in self.choice_set.all()
            ]

    def __str__(self):
        return self.type + ": " + self.text


class Choice(models.Model):
    text = models.CharField(verbose_name=_("Choice"), max_length=255)
    image = models.ImageField(verbose_name=_("Image"), blank=True, null=True)
    question = models.ForeignKey(Question, verbose_name=_("Question"), null=False, on_delete=models.CASCADE)

    class Meta:
        unique_together = ('text', 'question')
        verbose_name = _("Choice")
        verbose_name_plural = _("Choices")

    def clean(self):
        super().clean()

        if self.question.type == Question.RATE:
            raise ValidationError("No choice can be associated with a rate question")

    def __str__(self):
        return self.text


class Answer(models.Model):
    question = models.ForeignKey(Question, verbose_name=_("Question"), blank=False, null=False, on_delete=models.CASCADE)
    session = models.ForeignKey(Session, verbose_name=_("Session"), blank=False, null=False)

    class Meta:
        abstract = True
        verbose_name = _("Answer")
        verbose_name_plural = _("Answers")

    def clean(self):
        try:
            self.question
        except Question.DoesNotExist:
            raise ValidationError("The question does not exist or correspond to the answer type")


class UniqueAnswer(Answer):
    choice = models.ForeignKey(Choice, verbose_name=_("Choice"),  blank=False, on_delete=models.CASCADE)

    class Meta:
        verbose_name = _("Unique answer")
        verbose_name_plural = _("Unique answers")

    def clean(self, *args, **kwargs):
        super().clean()

        if self.question.type != Question.UNIQUE:
            raise ValidationError("Unique answer are valid only for unique choice questions")

        try:
            if self.choice.question != self.question:
                raise ValidationError("The choice must be assigned to the question")
        except Choice.DoesNotExist:
            raise ValidationError("The choice does not exist or correspond to the question")


class MultipleAnswer(Answer):
    choices = models.ManyToManyField(Choice, verbose_name=_("Choices"))
    # Note: many-to-many fields must be form-validated

    class Meta:
        verbose_name = _("Multiple answer")
        verbose_name_plural = _("Multiple answers")


class RankAnswer(Answer):
    choice = models.ForeignKey(Choice, verbose_name=_("Choice"),  blank=False, on_delete=models.CASCADE)
    rank = models.PositiveIntegerField(verbose_name=_("Ranking"))

    class Meta:
        verbose_name = _("Rank answer")
        verbose_name_plural = _("Rank answers")


class RateAnswer(Answer):
    rating = models.PositiveIntegerField(verbose_name=_("Rating"))

    class Meta:
        verbose_name = _("Rate answer")
        verbose_name_plural = _("Rate answers")

    def clean(self):
        super().clean()

        if self.question.type != Question.RATE:
            raise ValidationError(_("Rate answer are valid only for rate questions"))

        if self.rating > self.question.scale:
            raise ValidationError(
                _("The rating must not exceed the scale ({})")
                .format(self.question.scale)
            )


class Respondent(models.Model):
    first_name = models.CharField(verbose_name=_("First name"), max_length=64)
    last_name = models.CharField(verbose_name=_("Last name"), max_length=64)

    MALE = 'MALE'
    FEMALE = 'FEMA'
    UNDEFINED = 'NDEF'

    GENDERS = (
        (UNDEFINED, _("Undefined")),
        (MALE, _("Male")),
        (FEMALE, _("Female")),
    )
    gender = models.CharField(verbose_name=_("Gender"), max_length=4, choices=GENDERS, default=UNDEFINED)
    age = models.PositiveIntegerField(verbose_name=_("Age"), validators=[MaxValueValidator(120)], blank=True, null=True)
    session = models.ForeignKey(Session, verbose_name=_("Session"), null=False)

    class Meta:
        verbose_name = _("Respondent")
        verbose_name_plural = _("Respondents")

    def __str__(self):
        return f"{self.first_name} {self.last_name}"


class Comment(models.Model):
    author = models.ForeignKey(Respondent, verbose_name=_("Author"), null=True)
    text = models.TextField(verbose_name=_("Text"))
    session = models.ForeignKey(Session, verbose_name=_("Session"), null=False)
    date = models.DateTimeField(verbose_name=_("Date"), auto_now=True)

    class Meta:
        verbose_name = _("Comment")
        verbose_name_plural = _("Comments")

    def __str__(self):
        if len(self.text) > 16:
            return f"{self.text[:16]}..."
        else:
            return self.text
